"""
抽选结果导出模块
支持多种格式导出：Excel、CSV、JSON、PDF
"""

import json
import csv
import os
from datetime import datetime
from typing import List, Dict, Any

import pandas as pd
from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl未安装，Excel导出功能将不可用")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab未安装，PDF导出功能将不可用")

from .history import historys
from .stu import get_all_name


class ResultExporter:
    """抽选结果导出器"""
    
    def __init__(self):
        self.export_dir = "./exports"
        self.ensure_export_dir()
    
    def ensure_export_dir(self):
        """确保导出目录存在"""
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
            logger.info(f"创建导出目录: {self.export_dir}")
    
    def get_timestamp(self) -> str:
        """获取当前时间戳字符串"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def export_to_excel(self, history_data: List[Dict[str, Any]], filename: str = None) -> str:
        """导出到Excel格式"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl库未安装，无法导出Excel文件")
        
        if not filename:
            filename = f"抽选结果_{self.get_timestamp()}.xlsx"
        
        filepath = os.path.join(self.export_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "抽选结果"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 定义表头
        headers = ["序号", "抽选模式", "学生姓名", "学号", "抽选时间", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 填充数据
        for row_idx, history in enumerate(history_data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=self.get_mode_text(history.get("mode", 0)))
            
            student = history.get("student", {})
            ws.cell(row=row_idx, column=3, value=student.get("name", "未知"))
            ws.cell(row=row_idx, column=4, value=student.get("id", "未知"))
            
            # 格式化时间
            time_str = self.format_time(history.get("time"))
            ws.cell(row=row_idx, column=5, value=time_str)
            
            ws.cell(row=row_idx, column=6, value=history.get("note", ""))
        
        # 设置列宽
        column_widths = [8, 12, 15, 12, 20, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 添加边框
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(history_data)+1, max_col=6):
            for cell in row:
                cell.border = border
        
        wb.save(filepath)
        logger.info(f"Excel文件已导出: {filepath}")
        return filepath
    
    def export_to_csv(self, history_data: List[Dict[str, Any]], filename: str = None) -> str:
        """导出到CSV格式"""
        if not filename:
            filename = f"抽选结果_{self.get_timestamp()}.csv"
        
        filepath = os.path.join(self.export_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 写入表头
            headers = ["序号", "抽选模式", "学生姓名", "学号", "抽选时间", "备注"]
            writer.writerow(headers)
            
            # 写入数据
            for idx, history in enumerate(history_data, 1):
                row = [
                    idx,
                    self.get_mode_text(history.get("mode", 0)),
                    history.get("student", {}).get("name", "未知"),
                    history.get("student", {}).get("id", "未知"),
                    self.format_time(history.get("time")),
                    history.get("note", "")
                ]
                writer.writerow(row)
        
        logger.info(f"CSV文件已导出: {filepath}")
        return filepath
    
    def export_to_json(self, history_data: List[Dict[str, Any]], filename: str = None) -> str:
        """导出到JSON格式"""
        if not filename:
            filename = f"抽选结果_{self.get_timestamp()}.json"
        
        filepath = os.path.join(self.export_dir, filename)
        
        export_data = {
            "导出时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "总记录数": len(history_data),
            "学生名单": get_all_name(),
            "抽选结果": history_data
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, ensure_ascii=False, indent=2, default=str)
        
        logger.info(f"JSON文件已导出: {filepath}")
        return filepath
    
    def export_to_pdf(self, history_data: List[Dict[str, Any]], filename: str = None) -> str:
        """导出到PDF格式"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab库未安装，无法导出PDF文件")
        
        if not filename:
            filename = f"抽选结果_{self.get_timestamp()}.pdf"
        
        filepath = os.path.join(self.export_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # 居中
        )
        story.append(Paragraph("抽选结果报告", title_style))
        
        # 统计信息
        stats_text = f"""
        <para>
        <b>导出时间：</b>{datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}<br/>
        <b>总记录数：</b>{len(history_data)}条<br/>
        <b>学生总数：</b>{len(get_all_name())}人
        </para>
        """
        story.append(Paragraph(stats_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 表格数据
        table_data = [["序号", "抽选模式", "学生姓名", "学号", "抽选时间", "备注"]]
        
        for idx, history in enumerate(history_data, 1):
            row = [
                str(idx),
                self.get_mode_text(history.get("mode", 0)),
                history.get("student", {}).get("name", "未知"),
                str(history.get("student", {}).get("id", "未知")),
                self.format_time(history.get("time")),
                history.get("note", "")
            ]
            table_data.append(row)
        
        # 创建表格
        table = Table(table_data, colWidths=[0.5*inch, 1*inch, 1.2*inch, 1*inch, 1.5*inch, 2*inch])
        
        # 设置表格样式
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        logger.info(f"PDF文件已导出: {filepath}")
        return filepath
    
    def get_mode_text(self, mode: int) -> str:
        """获取抽选模式文本描述"""
        mode_map = {
            0: "个人抽选",
            1: "分组抽选",
            2: "权重抽选",
            3: "随机抽选"
        }
        return mode_map.get(mode, "未知模式")
    
    def format_time(self, time_obj) -> str:
        """格式化时间"""
        if isinstance(time_obj, str):
            try:
                # 尝试解析ISO格式时间字符串
                dt = datetime.fromisoformat(time_obj.replace('Z', '+00:00'))
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                return str(time_obj)
        elif hasattr(time_obj, 'strftime'):
            return time_obj.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return str(time_obj)
    
    def export_all_formats(self, history_data: List[Dict[str, Any]] = None) -> Dict[str, str]:
        """导出所有支持的格式"""
        if history_data is None:
            history_data = historys
        
        if not history_data:
            raise ValueError("没有可导出的历史记录")
        
        results = {}
        
        try:
            results["excel"] = self.export_to_excel(history_data)
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
        
        try:
            results["csv"] = self.export_to_csv(history_data)
        except Exception as e:
            logger.error(f"CSV导出失败: {e}")
        
        try:
            results["json"] = self.export_to_json(history_data)
        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
        
        try:
            results["pdf"] = self.export_to_pdf(history_data)
        except Exception as e:
            logger.error(f"PDF导出失败: {e}")
        
        return results


# 创建全局导出器实例
exporter = ResultExporter()


def export_recent_results(days: int = 7, format_type: str = "all") -> str:
    """导出最近的结果"""
    from datetime import datetime, timedelta
    
    cutoff_date = datetime.now() - timedelta(days=days)
    recent_history = []
    
    for history in historys:
        try:
            if isinstance(history.get("time"), str):
                history_time = datetime.fromisoformat(str(history.get("time")).replace('Z', '+00:00'))
            else:
                history_time = history.get("time")
            
            if history_time >= cutoff_date:
                recent_history.append(history)
        except:
            # 如果时间解析失败，包含所有记录
            recent_history.append(history)
    
    if not recent_history:
        raise ValueError(f"最近{days}天内没有抽选记录")
    
    if format_type == "all":
        return exporter.export_all_formats(recent_history)
    elif format_type == "excel":
        return exporter.export_to_excel(recent_history)
    elif format_type == "csv":
        return exporter.export_to_csv(recent_history)
    elif format_type == "json":
        return exporter.export_to_json(recent_history)
    elif format_type == "pdf":
        return exporter.export_to_pdf(recent_history)
    else:
        raise ValueError(f"不支持的格式: {format_type}")


def export_by_date_range(start_date: str, end_date: str, format_type: str = "all") -> str:
    """按日期范围导出结果"""
    from datetime import datetime
    
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        end_dt = end_dt.replace(hour=23, minute=59, second=59)
    except ValueError:
        raise ValueError("日期格式应为YYYY-MM-DD")
    
    filtered_history = []
    for history in historys:
        try:
            if isinstance(history.get("time"), str):
                history_time = datetime.fromisoformat(str(history.get("time")).replace('Z', '+00:00'))
            else:
                history_time = history.get("time")
            
            if start_dt <= history_time <= end_dt:
                filtered_history.append(history)
        except:
            continue
    
    if not filtered_history:
        raise ValueError(f"{start_date}到{end_date}期间没有抽选记录")
    
    if format_type == "all":
        return exporter.export_all_formats(filtered_history)
    else:
        return getattr(exporter, f"export_to_{format_type}")(filtered_history)


if __name__ == "__main__":
    # 测试导出功能
    try:
        results = exporter.export_all_formats()
        print("导出成功:", results)
    except Exception as e:
        print("导出失败:", str(e))